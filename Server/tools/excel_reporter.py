"""
Excel Reporter
==============
Generates a three-sheet .xlsx QC report using openpyxl.

Sheet 1 — Summary
  Total tests, passed, failed, pass rate %, generated timestamp.

Sheet 2 — Per-Testcase Detail
  Columns: Testcase ID | Name | Status | Failure Reason |
           Fix Suggestion | Description Gap | Description Gap Detail | AI Code Snippet
  Color coding: green rows = PASS, red rows = FAIL

Sheet 3 — QC Score
  Scores the project description on 4 parameters (1–5 each).
  Columns: Parameter | Critical | Score | Band | Max Grade | Justification | Author Action
  Overall grade and summary at the bottom.
"""
import logging
import os
from datetime import datetime
from typing import Dict, List, Optional

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# ── Style constants ───────────────────────────────────────────────────────────
_GREEN_FILL   = PatternFill("solid", fgColor="C6EFCE")
_RED_FILL     = PatternFill("solid", fgColor="FFC7CE")
_HEADER_FILL  = PatternFill("solid", fgColor="5C6BC0")
_LABEL_FILL   = PatternFill("solid", fgColor="E8EAF6")
_AMBER_FILL   = PatternFill("solid", fgColor="FFEB9C")
_ORANGE_FILL  = PatternFill("solid", fgColor="FFCC99")

# QC Score band colours (score 1→5)
_SCORE_FILL = {
    5: PatternFill("solid", fgColor="C6EFCE"),   # Excellent — green
    4: PatternFill("solid", fgColor="E2EFDA"),   # Good      — light green
    3: PatternFill("solid", fgColor="FFEB9C"),   # Acceptable — yellow
    2: PatternFill("solid", fgColor="FFCC99"),   # Needs Work — orange
    1: PatternFill("solid", fgColor="FFC7CE"),   # Reject     — red
}
_SCORE_FONT_COLOR = {
    5: "276221",
    4: "375623",
    3: "7D6608",
    2: "833C00",
    1: "9C0006",
}

_GREEN_FONT  = Font(color="276221")
_RED_FONT    = Font(color="9C0006")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_BOLD_FONT   = Font(bold=True)

_THIN = Side(style="thin")
_THIN_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)


# ── Public API ────────────────────────────────────────────────────────────────

def generate_excel_report(
    test_results:      List[Dict],
    test_summary:      Dict,
    failure_analysis:  List[Dict],
    report_path:       str,
    description_score: Optional[Dict] = None,
) -> str:
    """
    Build and save the Excel report.

    Args:
        test_results:      [{id, name, status, error_message}, ...]
        test_summary:      {total, passed, failed, pass_rate}
        failure_analysis:  [{id, name, category, description_gap,
                             description_gap_detail, implementation_detail,
                             fix_suggestion, code_snippet}, ...]
        report_path:       absolute or relative path for the .xlsx file
        description_score: output from DescriptionScorerAgent.score() or None

    Returns:
        Absolute path of the saved file.
    """
    wb = openpyxl.Workbook()

    _build_summary_sheet(wb, test_summary)
    _build_detail_sheet(wb, test_results, failure_analysis)
    _build_qc_score_sheet(wb, description_score)

    os.makedirs(os.path.dirname(os.path.abspath(report_path)), exist_ok=True)
    wb.save(report_path)
    abs_path = os.path.abspath(report_path)
    logger.info("Excel report saved → %s", abs_path)
    return abs_path


# ── Sheet builders ────────────────────────────────────────────────────────────

def _build_summary_sheet(wb: openpyxl.Workbook, summary: Dict) -> None:
    ws = wb.active
    ws.title = "Summary"

    ws.merge_cells("A1:B1")
    ws["A1"] = "QC Automation Report"
    ws["A1"].font      = Font(bold=True, size=16, color="5C6BC0")
    ws["A1"].alignment = _CENTER

    ws["A2"] = "Generated At"
    ws["A2"].font = _BOLD_FONT
    ws["B2"] = datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")

    ws["A3"] = ""

    rows = [
        ("Total Test Cases", summary.get("total",     0), None),
        ("Passed",           summary.get("passed",    0), _GREEN_FILL),
        ("Failed",           summary.get("failed",    0), _RED_FILL),
        ("Pass Rate",        f"{summary.get('pass_rate', 0.0):.1f}%",
         _GREEN_FILL if summary.get("pass_rate", 0) >= 70 else _RED_FILL),
    ]

    for i, (label, value, fill) in enumerate(rows, start=4):
        label_cell = ws.cell(row=i, column=1, value=label)
        label_cell.font      = _BOLD_FONT
        label_cell.fill      = _LABEL_FILL
        label_cell.border    = _THIN_BORDER
        label_cell.alignment = _LEFT

        val_cell = ws.cell(row=i, column=2, value=value)
        val_cell.border    = _THIN_BORDER
        val_cell.alignment = _CENTER
        if fill:
            val_cell.fill = fill
            val_cell.font = Font(
                bold=True,
                color="276221" if fill == _GREEN_FILL else "9C0006",
            )

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 28


def _build_detail_sheet(
    wb:               openpyxl.Workbook,
    test_results:     List[Dict],
    failure_analysis: List[Dict],
) -> None:
    ws = wb.create_sheet("Per-Testcase Detail")

    headers = [
        "Testcase ID",
        "Name",
        "Status",
        "Failure Reason",
        "Fix Suggestion",
        "Description Gap",
        "Description Gap Detail",
        "AI Generated Code Snippet",
    ]

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill      = _HEADER_FILL
        cell.font      = _HEADER_FONT
        cell.border    = _THIN_BORDER
        cell.alignment = _CENTER

    analysis_map: Dict[str, Dict] = {a["id"]: a for a in failure_analysis}

    for row_idx, result in enumerate(test_results, start=2):
        tc_id    = result["id"]
        status   = result["status"]
        is_pass  = status == "PASS"
        analysis = analysis_map.get(tc_id, {})

        row_fill       = _GREEN_FILL if is_pass else _RED_FILL
        row_font_color = "276221"    if is_pass else "9C0006"

        desc_gap        = "—"  if is_pass else ("Yes" if analysis.get("description_gap") else "No")
        desc_gap_detail = ""   if is_pass else analysis.get("description_gap_detail", "")
        failure_reason  = ""   if is_pass else analysis.get("implementation_detail",  "")
        fix_suggestion  = ""   if is_pass else analysis.get("fix_suggestion",         "")
        code_snippet    = ""   if is_pass else analysis.get("code_snippet",           "")

        row_data = [
            tc_id,
            result["name"],
            status,
            failure_reason,
            fix_suggestion,
            desc_gap,
            desc_gap_detail,
            code_snippet,
        ]

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill      = row_fill
            cell.font      = Font(color=row_font_color)
            cell.border    = _THIN_BORDER
            cell.alignment = _CENTER if col_idx <= 3 else _LEFT

    col_widths = [28, 32, 10, 50, 50, 16, 45, 55]
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"


def _build_qc_score_sheet(
    wb:                openpyxl.Workbook,
    description_score: Optional[Dict],
) -> None:
    ws = wb.create_sheet("QC Score")

    # ── Title ────────────────────────────────────────────────────────────────
    ws.merge_cells("A1:G1")
    ws["A1"] = "QC Score — Description Quality Assessment"
    ws["A1"].font      = Font(bold=True, size=14, color="5C6BC0")
    ws["A1"].alignment = _CENTER

    # ── Score band legend ────────────────────────────────────────────────────
    ws["A2"] = "Score"
    ws["B2"] = "Band"
    ws["C2"] = "Max Grade"
    ws["D2"] = "Meaning"
    ws["E2"] = "Author Action"
    for col in range(1, 6):
        ws.cell(row=2, column=col).font      = Font(bold=True, color="FFFFFF")
        ws.cell(row=2, column=col).fill      = PatternFill("solid", fgColor="37474F")
        ws.cell(row=2, column=col).border    = _THIN_BORDER
        ws.cell(row=2, column=col).alignment = _CENTER

    legend = [
        (5, "Excellent",   "A",      "No issues. Publish-ready.",                   "None required"),
        (4, "Good",        "A/B",    "Minor improvement possible but usable.",       "Optional improvement"),
        (3, "Acceptable",  "B/C",    "Correction required before Grade A.",          "Fix & re-review"),
        (2, "Needs Work",  "C/D",    "Significant issue. Return to author.",         "Return to author"),
        (1, "Reject",      "D only", "Complete failure. If ★ Critical → Grade D.",  "Reject / recreate"),
    ]
    for i, (score, band, grade, meaning, action) in enumerate(legend, start=3):
        fill = _SCORE_FILL[score]
        fc   = _SCORE_FONT_COLOR[score]
        data = [score, band, grade, meaning, action]
        for col_idx, val in enumerate(data, start=1):
            cell = ws.cell(row=i, column=col_idx, value=val)
            cell.fill      = fill
            cell.font      = Font(color=fc, bold=(col_idx == 1))
            cell.border    = _THIN_BORDER
            cell.alignment = _CENTER if col_idx <= 3 else _LEFT

    # ── Spacer ───────────────────────────────────────────────────────────────
    ws.row_dimensions[8].height = 10

    # ── Parameter scores table header ────────────────────────────────────────
    param_headers = [
        "Parameter",
        "Critical ★",
        "Score",
        "Band",
        "Max Grade",
        "Justification",
        "Author Action Required",
    ]
    for col_idx, h in enumerate(param_headers, start=1):
        cell = ws.cell(row=9, column=col_idx, value=h)
        cell.fill      = _HEADER_FILL
        cell.font      = _HEADER_FONT
        cell.border    = _THIN_BORDER
        cell.alignment = _CENTER

    if description_score and description_score.get("params"):
        params = description_score["params"]
    else:
        params = _placeholder_params()

    for row_idx, param in enumerate(params, start=10):
        score = param.get("score", 3)
        fill  = _SCORE_FILL.get(score, _AMBER_FILL)
        fc    = _SCORE_FONT_COLOR.get(score, "7D6608")

        row_data = [
            param["name"],
            "★ Critical" if param["critical"] else "—",
            score,
            param["band"],
            param["max_grade"],
            param.get("justification", ""),
            param.get("author_action", param.get("action", "")),
        ]
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill      = fill
            cell.font      = Font(color=fc, bold=(col_idx == 3))
            cell.border    = _THIN_BORDER
            cell.alignment = _CENTER if col_idx in (2, 3, 4, 5) else _LEFT

    # ── Overall grade ────────────────────────────────────────────────────────
    overall_row = 10 + len(params) + 1
    grade  = (description_score or {}).get("overall_grade", "?")
    scores = [(description_score or {}).get(f"p{i}_score", 3) for i in range(1, 5)]
    total  = sum(scores)

    grade_fill = {
        "A": PatternFill("solid", fgColor="C6EFCE"),
        "B": PatternFill("solid", fgColor="E2EFDA"),
        "C": PatternFill("solid", fgColor="FFEB9C"),
        "D": PatternFill("solid", fgColor="FFC7CE"),
    }.get(grade, _AMBER_FILL)
    grade_fc = {
        "A": "276221", "B": "375623", "C": "7D6608", "D": "9C0006",
    }.get(grade, "7D6608")

    ws.merge_cells(f"A{overall_row}:B{overall_row}")
    ws[f"A{overall_row}"] = "Overall QC Grade"
    ws[f"A{overall_row}"].font      = Font(bold=True, size=12)
    ws[f"A{overall_row}"].fill      = _LABEL_FILL
    ws[f"A{overall_row}"].border    = _THIN_BORDER
    ws[f"A{overall_row}"].alignment = _CENTER

    ws[f"C{overall_row}"] = grade
    ws[f"C{overall_row}"].font      = Font(bold=True, size=14, color=grade_fc)
    ws[f"C{overall_row}"].fill      = grade_fill
    ws[f"C{overall_row}"].border    = _THIN_BORDER
    ws[f"C{overall_row}"].alignment = _CENTER

    ws[f"D{overall_row}"] = f"Total Score: {total}/20"
    ws[f"D{overall_row}"].font      = Font(bold=True)
    ws[f"D{overall_row}"].fill      = grade_fill
    ws[f"D{overall_row}"].border    = _THIN_BORDER
    ws[f"D{overall_row}"].alignment = _CENTER

    summary_row = overall_row + 1
    ws.merge_cells(f"A{summary_row}:G{summary_row}")
    ws[f"A{summary_row}"] = (description_score or {}).get(
        "overall_summary", "Run the pipeline to generate scores."
    )
    ws[f"A{summary_row}"].font      = Font(italic=True, color="444444")
    ws[f"A{summary_row}"].fill      = _LABEL_FILL
    ws[f"A{summary_row}"].border    = _THIN_BORDER
    ws[f"A{summary_row}"].alignment = _LEFT

    # ── Column widths ────────────────────────────────────────────────────────
    col_widths = [38, 14, 8, 16, 12, 55, 45]
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.row_dimensions[9].height = 28


def _placeholder_params() -> list:
    return [
        {"name": "Milestone / Deliverable Def.",      "critical": False, "score": 3, "band": "Acceptable", "max_grade": "B/C", "justification": "Not scored — run pipeline.", "author_action": "Manual review"},
        {"name": "Expected Output Clarity",            "critical": False, "score": 3, "band": "Acceptable", "max_grade": "B/C", "justification": "Not scored — run pipeline.", "author_action": "Manual review"},
        {"name": "Execution Instruction Clarity ★",   "critical": True,  "score": 3, "band": "Acceptable", "max_grade": "B/C", "justification": "Not scored — run pipeline.", "author_action": "Manual review"},
        {"name": "Difficulty vs Effort Calibration",  "critical": False, "score": 3, "band": "Acceptable", "max_grade": "B/C", "justification": "Not scored — run pipeline.", "author_action": "Manual review"},
    ]
